import re
import pandas as pd
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np

def extract_numbers_from_text(text):
    """Extracts numbers from a given text using regular expressions."""
    numbers = re.findall(r':\s*(-?\d+\.*\d*)\n', text)
    return [float(num) for num in numbers]

def create_graph(numbers, name):
    """Creates a graph from the extracted numbers."""
    plt.plot(numbers)
    plt.xlabel('Runtime')
    plt.ylabel('Value')
    plt.title(name)
    plt.grid(True)
    plt.show()


# Read text file
with open('GYRO - Copy.txt', 'r') as file:
    text = file.read()

# Extract numbers
numbers = extract_numbers_from_text(text)

# Create Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Numbers"

# Write numbers to Excel sheet
row = 1
column = 1
for number in numbers:
    ws.cell(row=row, column=column, value=number)
    if column == 3:
        row += 1
        column = 1
    else:
        column += 1

roll = []
pitch = []
yaw = []
count = 0
for col in ws.iter_cols(min_row=57000, max_col=3, max_row = 59473, values_only=True):
    count += 1
    if count == 1:
        for value in col:
            if abs(value) < 180:
                yaw.append(value)
            elif abs(value) < 500:
                yaw.append(value-360)
    elif count == 2:
        for value in col:
            if abs(value) < 500:
                roll.append(value)
    else:
        for value in col:
            if abs(value) < 500:
                pitch.append(value)

yaw = [item * -1 for item in yaw]

df = pd.read_csv('FlightData6D.tsv',sep='\t',skiprows=11)
gimbal_pos = df[['Gimbal1 X','Y','Z']].to_numpy()
gimbal_rpy = df[['Roll','Pitch','Yaw']].to_numpy()
drone_pos = df[['M2 X','Y.2','Z.2']].to_numpy()
drone_rpy = df[['Roll.2','Pitch.2','Yaw.2']].to_numpy()

SDxAxis = np.linspace(0, 12000, num = 2474)


labels = ['motion capture roll', 'motion capture pitch', 'motion capture yaw']
SDlabels = ["SD card roll", "SD card pitch", "SD card yaw"]
axis = [roll, pitch, yaw]
for i in range(3):
    plt.subplot(1,3,i+1)
    plt.plot(drone_rpy[:,i], label=f'drone-{labels[i]}')
    plt.plot(SDxAxis,axis[i], label = f'drone-{SDlabels[i]}')
    plt.legend()

plt.show()

#create_graph(yaw, "Yaw")
#create_graph(roll, "Roll")
#create_graph(pitch, "Pitch")


# Save Excel file
wb.save("output.xlsx")
print("Numbers extracted and saved to output.xlsx")

